
from cgitb import text
import re
import openpyxl
import textstat



def reading_data_from_xlsx_file(file:str):
    dataframe = openpyxl.load_workbook(file)
    sheet_obj = dataframe.active
    text = []
    for i in range(2, sheet_obj.max_row+1):
        key = sheet_obj.cell(row=i, column=1)
        title = sheet_obj.cell(row=i, column=2)
        description = sheet_obj.cell(row=i, column=3)
        response = sheet_obj.cell(row=i, column=4)
        conclusion = sheet_obj.cell(row=i, column=5)
        text.append({'key':key.value, 'title':title.value, 'description':description.value, 'response':response.value, 'conclusion':conclusion.value})
    return text
        